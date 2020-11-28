
"""
Author: Mario Uriel Romero Martínez
Contact: marioma3@cisco.com
Purpose: To collect all information regarding control sessions in the SD-WAN fabric
"""

import json
import requests
from authentication import *
from utils import *
from getpass import getpass
from openpyxl import Workbook
from openpyxl import load_workbook

#base_url="https://vmanage-00000.viptela.net/"

vmng= input("vManage URL(without https://) or IP: ")
base_url=f"https://{vmng}"

user = input("Username: ")
passw = getpass("Password: ")

def get_deviceinfo(jsessionid,token):

    listaux=[]
    dictaux={}
    
    if token is not None:
        header = {'Content-Type': "application/json", 'Cookie': jsessionid, 'X-XSRF-TOKEN': token}
    else:
        header = {'Content-Type': "application/json", 'Cookie': jsessionid}

    device_url = "dataservice/device"

    response=requests.get(url=f"{base_url}{device_url}",headers=header,verify=False)
    if response.status_code==200:
        print("Getting device info...")
        devices=response.json()["data"]
        for device in devices:
            listaux.append({"systemip":device["system-ip"],"host-name":device["host-name"],"uuid":device["uuid"]})


        dictaux["data"]=listaux
        savefile("vedges_info.json",dictaux)
         
    else:
        print("Error ",device_response.status_code,device_response.text)



def get_concon(jsessionid,token,jsonfile):

    book = Workbook()
    sheet = book.active

    if token is not None:
        header = {'Content-Type': "application/json", 'Cookie': jsessionid, 'X-XSRF-TOKEN': token}
    else:
        header = {'Content-Type': "application/json", 'Cookie': jsessionid}
    
    
    sheet.cell(row=1, column=1).value = "host-name"
    sheet.cell(row=1, column=2).value = "system-ip"
    sheet.cell(row=1, column=3).value = "uuid"
    sheet.cell(row=1, column=4).value = "local color"
    sheet.cell(row=1, column=4).value = "peer type"
    sheet.cell(row=1, column=6).value = "peer system ip"
    sheet.cell(row=1, column=7).value = "peer site-id"
    sheet.cell(row=1, column=8).value = "peer private ip"
    sheet.cell(row=1, column=9).value = "peer public ip"
    sheet.cell(row=1, column=10).value = "remote color"
    sheet.cell(row=1, column=11).value = "state"
    sheet.cell(row=1, column=12).value = "uptime"

    with open(jsonfile,'r') as f:
        jparse=f.read()
        jfile=json.loads(jparse)["data"]


    rows=2

    for device in jfile:
        ccurl=f"dataservice/device/control/connections?deviceId={device['systemip']}"
        response=requests.get(url=f"{base_url}{ccurl}",headers=header,verify=False)

        if response.status_code==200:

            sheet.cell(row=rows, column=1).value = device["host-name"]
            sheet.cell(row=rows, column=2).value = device["systemip"]
            sheet.cell(row=rows, column=3).value = device["uuid"]            

            ccdata=response.json()["data"]
            print(f"Getting control connections for {device['host-name']}")
            for cc in ccdata:
                sheet.cell(row=rows, column=4).value = cc["local-color"]
                sheet.cell(row=rows, column=5).value = cc["peer-type"]
                sheet.cell(row=rows, column=6).value = cc["system-ip"]
                sheet.cell(row=rows, column=7).value = cc["site-id"]
                sheet.cell(row=rows, column=8).value = cc["private-ip"]
                sheet.cell(row=rows, column=9).value = cc["public-ip"]
                sheet.cell(row=rows, column=10).value = cc["remote-color"]
                sheet.cell(row=rows, column=11).value = cc["state"]
                sheet.cell(row=rows, column=12).value = convert_epochtime(cc["uptime-date"])

                rows+=1
        else:
            print("Error ",response.status_code,response.text)
            
    print("Saving spreadsheet")
    book.save('Control_Connections.xlsx')


def get_conlocwanint(jsessionid,token,jsonfile,excelfile):

    if token is not None:
        header = {'Content-Type': "application/json", 'Cookie': jsessionid, 'X-XSRF-TOKEN': token}
    else:
        header = {'Content-Type': "application/json", 'Cookie': jsessionid}
    
    workbook = load_workbook(filename=excelfile)
    sheet = workbook["Sheet"]
    sheet2 = workbook.create_sheet("Sheet2")


    sheet2['A' + str(1)].value = "host-name"
    sheet2['B' + str(1)].value = "vdevice system-ip"
    sheet2['C' + str(1)].value = "color"
    sheet2['D' + str(1)].value = "interface"
    sheet2['E' + str(1)].value = "admin-state"
    sheet2['F' + str(1)].value = "operation-state"
    sheet2['G' + str(1)].value = "num-vsmarts"
    sheet2['H' + str(1)].value = "num-vmanages"

    with open(jsonfile,'r') as f:
        jparse=f.read()
        jfile=json.loads(jparse)["data"]

    row = 2

    for device in jfile:

        clwi_url=f"dataservice/device/control/waninterface?deviceId={device['systemip']}"

        response= requests.get(url=f"{base_url}{clwi_url}",headers=header,verify=False)
        if response.status_code==200:
            print(f"Getting Control Local WAN Interfce info for {device['host-name']}")
            wanc=response.json()["data"]

            sheet2['A' + str(row)].value = device["host-name"]

            for w in wanc:
                
                sheet2['B' + str(row)].value = w["vdevice-name"]
                sheet2['C' + str(row)].value = w["color"]
                sheet2['D' + str(row)].value = w["interface"]
                sheet2['E' + str(row)].value = w["admin-state"]
                sheet2['F' + str(row)].value = w["operation-state"]
                sheet2['G' + str(row)].value = w["num-vsmarts"]
                sheet2['H' + str(row)].value = w["num-vmanages"]
                
                row+=1
    
        else:
            print("Error ",response.status_code,response.text)

    print("Saving spreadsheet")
    workbook.save('Control_Connections.xlsx')


                
if __name__ == "__main__":
    jsessionid=get_jsessionid(base_url,user,passw)
    token=get_token(jsessionid,base_url)
    get_deviceinfo(jsessionid,token)
    get_concon(jsessionid,token,"vedges_info.json")
    get_conlocwanint(jsessionid,token,"vedges_info.json","Control_Connections.xlsx")









        







                


        




