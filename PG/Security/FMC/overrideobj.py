
"""
Author: Mario Uriel Romero Martínez
Organization:CX PS
Contact: marioma3@cisco.com
Description: Python script to get override objects (Networks/Hosts) and create a spreadsheet report from Firepower Management Center 
            

"""

import json
import requests
import sys
import datetime
from openpyxl import Workbook
from getpass import getpass

requests.packages.urllib3.disable_warnings()

book = Workbook()
sheet = book.active

   
   
def authenticate(username,password):
    dictauth={}   
    headers = {'Content-Type': 'application/json'}
    auth_url = "/api/fmc_platform/v1/auth/generatetoken"
    username = user
    password = passw

    try:
        login_response=requests.post(f"{fmc_url}{auth_url}",auth=(username,password),verify=False)
        print(login_response.status_code)
        resp_headers=login_response.headers
        token = resp_headers.get('X-auth-access-token', default=None)
        domainid = resp_headers.get('DOMAIN_UUID')
        #headers['X-auth-access-token'] = token
        dictauth.update({"Token":token,"DomainId":domainid})
        
    except Exception as e:
        print ("Error in authentication ", e)

    
    #dictauth["token"]=token
    #dictauth["domainid"]=domainid

    return dictauth

def get_devices(token,domainid):

    headers = {'Content-Type': 'application/json'}
    headers['X-auth-access-token'] = token

    #Getting the UUID for all devices attached to the FMC
    dev_url=f"/api/fmc_config/v1/domain/{domainid}/devices/devicerecords"
    dev_response=requests.get(url=f"{fmc_url}{dev_url}",headers=headers,verify=False)
    if dev_response.status_code==200:
        #dev_response=dev_response.json()["items"][0]["id"]
        #print(dev_response)
        devices=dev_response.json()["items"]
        listdevices=[]
        for item in devices:
            listdevices.append({"DeviceName":item["name"],"DeviceId":item["id"]})
    else:
        print("Error trying to collect device information... ",dev_response.status_code,dev_response.text)

        #print(listdevices)
    return listdevices

def get_obj_id(token,domainid,dev,ot):

    headers = {'Content-Type': 'application/json'}
    headers['X-auth-access-token'] = token


    #ovr_url=f"/api/fmc_config/v1/domain/{domainid}/object/networks/005056BB-0B24-0ed3-0000-893353264061/overrides"
    #dev_response contains the DevId.
    ##for dev in listdevices:

    ovr_url=f"/api/fmc_config/v1/domain/{domainid}/object/{ot}?overrideTargetId={dev}"
    #ovr_url = f"/api/fmc_config/v1/domain/{domainid}/object/networks"
    #ovr_url = f"/api/fmc_config/v1/domain/{domainid}/object/hosts?expanded=true&overrideTargetId"
    #ovr_url = f"/api/fmc_config/v1/domain/{domainid}/object/hosts/{dev_response}/overrides"

    ovr_response=requests.get(url=f"{fmc_url}{ovr_url}", headers=headers, verify=False)

    listobj=[]
    if ovr_response.status_code==200:
        ovr_response=ovr_response.json()
        #print(json.dumps(ovr_response,indent=4))
        for dictio in ovr_response["items"]:
            listobj.append(dictio["id"])

        #print(listobj)

    else:
        print("Error trying to collect object Id... fmcrestapisandbox.cisco.com",ovr_response.status_code,ovr_response.text)
    
    return listobj

def get_obj_data(token, domainid,listobj,ot):

    headers = {'Content-Type': 'application/json'}
    headers['X-auth-access-token'] = token

    dictobj={}
    listobjdat=[]
    for obj in listobj:   
        obj_url=f"/api/fmc_config/v1/domain/{domainid}/object/{ot}/{obj}/overrides"
        obj_response=requests.get(f"{fmc_url}{obj_url}",headers=headers,verify=False)
        if obj_response.status_code==200:
            objdetail=obj_response.json()["items"]
            for item in objdetail:
                listobjdat.append({"Type":item["type"],"Value":item["value"],"Device":item["overrides"]["target"]["name"]})
        
        else:
            print("Error trying to collect object data...",obj_response.status_code,obj_response.text)
            
    #dictobj["data"]=listobjdet
    
    #print(json.dumps(dictobj,indent=4))
    return listobjdat

def print_ovrobj(username,passw,ot):
    dictauth={}
    dictdevobj={}
    listobjdata=[]
    listdevobj=[]
    listdev=[]
    listobjid=[]
    
    #Getting all devices attached to FMC
    listdev=get_devices(token,domainid)
    #For each device, look for the override obj ids
    rdev=1
    for dev in listdev:
        
        sheet.cell(row=rdev, column=1).value =dev["DeviceName"]
        #Get the obj ids
        listobjid=get_obj_id(token,domainid,dev["DeviceId"],ot)
        #Get the override obj info for all obj ids
        listobjdata=get_obj_data(token,domainid,listobjid,ot)
        listdevobj.append({"Device":dev["DeviceName"],"Override_objects":listobjdata})
        for dobj in listobjdata:
            objdr=rdev
            for k,v in dobj.items():
                
                sheet.cell(row=objdr,column=2).value=k
                sheet.cell(row=objdr,column=3).value=v
                objdr+=1
            rdev=objdr+1

    dictdevobj["data"]=listdevobj
    print(json.dumps(dictdevobj,indent=4))
    book.save(f'Override_{ot}_{datetime.datetime.now().date()}.xlsx')    


    

    
if __name__ == "__main__":

    
    fmc= input("FMC URL(without https://) or IP: ")
    fmc_url=f"https://{fmc}"
    #fmc_url = "https://fmcrestapisandbox.cisco.com"

    user = input("Username: ")
    passw = getpass("Password: ")
    dictauth=authenticate(user,passw)
    token=dictauth["Token"]
    domainid=dictauth["DomainId"]

    if token == None:
        sys.exit("Authentication error...exiting from code")

    else:
        print("Authetincation was succesful!")

        while True:
            objtype = input("Select Override Type: 1-Host, 2-Networks, 3.Exit: ")

            if int(objtype)==1:
                ot="hosts"
            elif int(objtype)==2:
                ot="networks"
            
            elif int(objtype)==3:
                sys.exit("Exit option selected...exiting from code")

            else:
                sys.exit("Invalid option...exiting from code")
    
            print_ovrobj(user,passw,ot)



