import json
from utils import savefile
import requests
import sys
from datetime import date
from openpyxl import Workbook
from getpass import getpass

"""
Author: Mario Uriel Romero Martínez
Organization:CX BCS SD-WAN
Description: Python script to generate a report for Network Group objects from Firepower Management Center 
"""

requests.packages.urllib3.disable_warnings()

book = Workbook()
sheet = book.active

   
   
def authenticate(username,password,fmc_url):
    dictauth={}   
    headers = {'Content-Type': 'application/json'}
    auth_url = "/api/fmc_platform/v1/auth/generatetoken"

    try:
        login_response=requests.post(f"{fmc_url}{auth_url}",auth=(username,password),verify=False)
        print(login_response.status_code)
        resp_headers=login_response.headers
        token = resp_headers.get('X-auth-access-token', default=None)
        domainid = resp_headers.get('DOMAIN_UUID')
        #headers['X-auth-access-token'] = token
        dictauth.update({"Token":token,"DomainId":domainid})
        
    except Exception as e:
        print ("Error in authentication ", e)

    return dictauth

def get_nobjg(token,fmc_url,domainid):

    listobjdat=[]
    headers = {'Content-Type': 'application/json'}
    headers['X-auth-access-token'] = token
    
    nog_url=f"/api/fmc_config/v1/domain/{domainid}/object/networkgroups"
    response=requests.get(url=f"{fmc_url}{nog_url}",headers=headers,verify=False)
    if response.status_code==200:
        nog_data=response.json()["items"]
        #print(json.dumps(nog_data,indent=4))
        for objg in nog_data:
            listobjdat.append(objg["id"])
    else:
        ("Error",response.status_code,response.text)
    
    return listobjdat

def get_objinfo(token,fmc_url,domainid,listnog):

    headers = {'Content-Type': 'application/json'}
    headers['X-auth-access-token'] = token

    listaux=[]
    dictaux={}

    for nogid in listnog:
        #nobj_url=f"/api/fmc_config/v1/domain/{domainid}/object/networks/{nogid}"
        nobj_url=f"/api/fmc_config/v1/domain/{domainid}/object/networkgroups/{nogid}"
        response=requests.get(url=f"{fmc_url}{nobj_url}",headers=headers,verify=False)
        if response.status_code==200:
            obj_data=response.json()
            #print(json.dumps(obj_data,indent=4))
            if "objects" in obj_data:
                value=obj_data["objects"]
            else:
                value=None
            listaux.append({"Name":obj_data["name"],"Value":value,"Type":obj_data["type"]})
            

        else:
            print("Error",response.status_code,response.text)
    dictaux["data"]=listaux
    savefile("netgrobj.json",dictaux)

def writetoexcel(jsonfile):
    today = date.today()
    book = Workbook()
    sheet = book.active

    with open(jsonfile,'r') as f:
        jparse=f.read()
        objdata=json.loads(jparse)["data"]
    
    sheet.cell(row=1, column=1).value="Name"
    sheet.cell(row=1, column=2).value="Value"
    sheet.cell(row=1, column=3).value="Type"

    rows=2
    for nog in objdata:
        if nog["Value"]!=None:
            for v in nog["Value"]:
                sheet.cell(row=rows, column=1).value=nog["Name"]
                sheet.cell(row=rows, column=3).value=nog["Type"]
                sheet.cell(row=rows, column=2).value=v["name"]
                rows+=1
        else:
            rows+=1

    book.save(f"NetworkObjectGroup{today}.xlsx")
    book.close()




# if __name__ == "__main__":
#     fmc_url = "https://fmcrestapisandbox.cisco.com"
#     username="marioma3"
#     password="PxdEvGXw"
#     listobj=[]
    #dictauth=authenticate(username,password,fmc_url)
    #token=dictauth["Token"]
    #domainid=dictauth["DomainId"]
    #listobj=get_nobjg(token,fmc_url,domainid)
    #listobj=["005056BB-0B24-0ed3-0000-893353505084"]
    #print(listobj)
    #get_objinfo(token,fmc_url,domainid,listobj)
    #writetoexcel("netgrobj.json")





